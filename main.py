import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Step 1: Read both Excel files
df_jan = pd.read_excel("data/raw_sales_january.xlsx")
df_feb = pd.read_excel("data/raw_sales_february.xlsx")

# Step 2: Combine them into one DataFrame
df = pd.concat([df_jan, df_feb], ignore_index=True)

# Step 3: Clean the data
df.dropna(subset=["Customer", "Amount"], inplace=True)
df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
df = df.dropna(subset=["Amount"])

# Step 4: Group by Customer and summarize
summary = df.groupby("Customer", as_index=False)["Amount"].sum()
summary.columns = ["Customer", "Total Spent"]

# Step 5: Save the summary to Excel
os.makedirs("output", exist_ok=True)
output_path = "output/cleaned_summary_report.xlsx"
summary.to_excel(output_path, index=False)

# Step 6: Format the header using openpyxl
wb = load_workbook(output_path)
ws = wb.active
for cell in ws[1]:
    cell.font = Font(bold=True)
wb.save(output_path)

print("✅ Cleaned summary report saved at:", output_path)
